from Employee import Employee
from Order import Order
import openpyxl
from Worker import Worker
import tkinter as tk


class Admin(Employee):
    def __init__(self, pk, name, designation, possession):
        super().__init__(pk, name, designation, possession)

    
    def giveOrder(self, workerId, machineId, productId, toWork, toDo):
        """Giving order to the worker"""
        order = Order(workerId, self.getPrimaryKey(), machineId, productId, toWork, toDo)
        worker = Admin.findWorker(workerId)
        worker.getOrder(order)

    def save(self):
        """For saving object ot the excel file"""
        WB = openpyxl.load_workbook("project.xlsx")
        admins = WB["admins"]
        lastrow = admins.max_row + 1
        self.lastrow = lastrow
        admins[f"A{lastrow}"].value = self.getPrimaryKey()
        admins[f"B{lastrow}"].value = self.getName()
        admins[f"C{lastrow}"].value = self.getDesignation()
        admins[f"D{lastrow}"].value = self.getPossessions()
        WB.save("project.xlsx")
        
    def giveReport(frame, r):
        """
        Report of all the objects created
        frame => frame on which information is required to be rendered
        """
        WB = openpyxl.load_workbook("project.xlsx")
        admins = WB["admins"]
        MAXROW_admin = admins.max_row
        for i in range(2, MAXROW_admin+1):
            tk.Label(frame, text=admins[f"A{i}"].value, font=("Times New Roman", 13)).grid(row=r, column=0)
            tk.Label(frame, text=admins[f"B{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=1)
            tk.Label(frame, text=admins[f"C{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=2)
            tk.Label(frame, text=admins[f"D{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=3)
            r = r + 1
        WB.save("project.xlsx")
        return r
        

    def findWorker(id):
        # find the worker object in file
        WB = openpyxl.load_workbook("project.xlsx")
        workers = WB["workers"]
        MAXROW_admin = workers.max_row
        for i in range(1, MAXROW_admin+1):
            if str(workers[f"A{i}"].value) == str(id):
                return Worker(id, workers[f"B{i}"].value, workers[f"C{i}"].value, workers[f"D{i}"].value)
        WB.save("project.xlsx")

    def findAdmin(id):
        # find the admin object in excel file
        WB = openpyxl.load_workbook("project.xlsx")
        admins = WB["admins"]
        MAXROW_admin = admins.max_row
        for i in range(1, MAXROW_admin+1):
            if str(admins[f"A{i}"].value) == str(id):
                return Admin(id, admins[f"B{i}"].value, admins[f"C{i}"].value, admins[f"D{i}"].value)
        WB.save("project.xlsx")
    
    def isExist(id):
        # find the admin object in excel file
        WB = openpyxl.load_workbook("project.xlsx")
        admins = WB["admins"]
        MAXROW_admin = admins.max_row
        for i in range(1, MAXROW_admin+1):
            if str(admins[f"A{i}"].value) == str(id):
                return True
        WB.save("project.xlsx")
        return False

    def seeOrders(frame):
        WB = openpyxl.load_workbook("project.xlsx")
        works = WB["works"]
        MAXROW_admin = works.max_row
        r = 2
        for i in range(2, MAXROW_admin+1):
            tk.Label(frame, text=f"{works[f'A{i}'].value} orderd {works[f'B{i}'].value} to {works[f'E{i}'].value} and {works[f'B{i}'].value} used {works[f'C{i}'].value} to {works[f'F{i}'].value} and make {works[f'D{i}'].value}", font=("Times New Roman", 13)).grid(row=r, column=0)
            # tk.Label(frame, text=works[f"B{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=1)
            # tk.Label(frame, text=works[f"C{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=2)
            # tk.Label(frame, text=works[f"D{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=3)
            r = r + 1
        WB.save("project.xlsx")

# if __name__ == "__main__":
#     ad = Admin.findAdmin(1)
#     ad.giveOrder(2, 3, 4, "design software", "programmed")
