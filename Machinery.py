from Property import Property
import openpyxl
from Product import Product
import tkinter as tk



class Machinery(Property):
    def __init__(self, pk, name, type):
        super().__init__(pk, name)
        self.__type = type


    def save(self):
        """For saving object ot the excel file"""
        WB = openpyxl.load_workbook("project.xlsx")
        machine = WB["machines"]
        lastrow = machine.max_row + 1
        self.lastrow = lastrow
        machine[f"A{lastrow}"].value = self.getPrimaryKey()
        machine[f"B{lastrow}"].value = self.getName()
        machine[f"C{lastrow}"].value = self.__type
        WB.save("project.xlsx")

    def getType(self):
        """Getter for the type of the machine"""
        return self.__type

    def prepareProduct(self, order):
        from Admin import Admin
        from Worker import Worker
        WB = openpyxl.load_workbook("project.xlsx")
        works = WB["works"]
        product = Machinery.findProduct(order.getProduct())
        tempRow = works.max_row + 1
        works[f"A{tempRow}"].value = Admin.findAdmin(order.getAdmin()).getName()
        works[f"B{tempRow}"].value = Admin.findWorker(order.getWorker()).getName()
        works[f"C{tempRow}"].value = Worker.findMachine(order.getMachine()).getName()
        works[f"D{tempRow}"].value = Machinery.findProduct(product.getPrimaryKey()).getName()
        works[f"E{tempRow}"].value = order.getWork()
        works[f"F{tempRow}"].value = order.getDo()
        WB.save("project.xlsx")

    def findProduct(id):
        WB = openpyxl.load_workbook("project.xlsx")
        products = WB["products"]
        MAXROW_product = products.max_row
        for i in range(1, MAXROW_product+1):
            if str(products[f"A{i}"].value) == str(id):
                return Product(id, products[f"B{i}"].value, products[f"C{i}"].value)
        WB.save("project.xlsx")

    def giveReport(frame, r):
        """
        Report of all the objects created
        frame => frame on which information is required to be rendered
        """
        WB = openpyxl.load_workbook("project.xlsx")
        machines = WB["machines"]
        MAXROW_admin = machines.max_row
        for i in range(2, MAXROW_admin+1):
            tk.Label(frame, text=machines[f"A{i}"].value, font=("Times New Roman", 13)).grid(row=r, column=0)
            tk.Label(frame, text=machines[f"B{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=1)
            tk.Label(frame, text=machines[f"C{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=2)
            r = r + 1
        WB.save("project.xlsx")
        return r

    def isExist(id):
        WB = openpyxl.load_workbook("project.xlsx")
        machines = WB["machines"]
        MAXROW_machine = machines.max_row
        for i in range(1, MAXROW_machine+1):
            if str(machines[f"A{i}"].value) == str(id):
                return True
        WB.save("project.xlsx")
        return False
