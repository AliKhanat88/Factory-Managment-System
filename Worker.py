from Employee import Employee

import openpyxl
import tkinter as tk


class Worker(Employee):
    def __init__(self, pk, name, designation, possession):
        super().__init__(pk, name, designation, possession)

    def save(self):
        """For saving object ot the excel file"""
        WB = openpyxl.load_workbook("project.xlsx")
        workers = WB["workers"]
        lastrow = workers.max_row + 1
        self.lastrow = lastrow
        workers[f"A{lastrow}"].value = self.getPrimaryKey()
        workers[f"B{lastrow}"].value = self.getName()
        workers[f"C{lastrow}"].value = self.getDesignation()
        workers[f"D{lastrow}"].value = self.getPossessions()
        WB.save("project.xlsx")
    
    def getOrder(self, order):
        """For getting order from the Admin"""
        machineId = order.getMachine()
        machine = Worker.findMachine(machineId)
        machine.prepareProduct(order)

    def getWork(self):
        """Getter for the work"""
        return self.__work

    def giveReport(frame, r):
        """
        Report of all the objects created
        frame => frame on which information is required to be rendered
        """
        WB = openpyxl.load_workbook("project.xlsx")
        workers = WB["workers"]
        MAXROW_admin = workers.max_row
        for i in range(2, MAXROW_admin+1):
            tk.Label(frame, text=workers[f"A{i}"].value, font=("Times New Roman", 13)).grid(row=r, column=0)
            tk.Label(frame, text=workers[f"B{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=1)
            tk.Label(frame, text=workers[f"C{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=2)
            tk.Label(frame, text=workers[f"D{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=3)
            r = r + 1
        WB.save("project.xlsx")
        return r

    def isExist(id):
        WB = openpyxl.load_workbook("project.xlsx")
        workers = WB["workers"]
        MAXROW_admin = workers.max_row
        for i in range(1, MAXROW_admin+1):
            if str(workers[f"A{i}"].value) == str(id):
                return True
        WB.save("project.xlsx")
        return False

    def findMachine(id):
        """Find machine object in the file"""
        from Machinery import Machinery
        WB = openpyxl.load_workbook("project.xlsx")
        machines = WB["machines"]
        MAXROW_machine = machines.max_row
        for i in range(1, MAXROW_machine+1):
            if str(machines[f"A{i}"].value) == str(id):
                return Machinery(id, machines[f"B{i}"].value, machines[f"C{i}"].value)
        WB.save("project.xlsx")

