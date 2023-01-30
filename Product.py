from Property import Property
import openpyxl
import tkinter as tk

class Product(Property):
    def __init__(self, pk, name, amount):
        super().__init__(pk, name)
        self.__amount = amount

        
    def save(self):
        """For saving object ot the excel file"""
        WB = openpyxl.load_workbook("project.xlsx")
        products = WB["products"]
        MAXROW = products.max_row + 1
        products[f"A{MAXROW}"].value = self.getPrimaryKey()
        products[f"B{MAXROW}"].value = self.getName()
        products[f"C{MAXROW}"].value = self.__amount
        WB.save("project.xlsx")

    def getAmount(self):
        """Getter for the amount of product"""
        return self.__amount


    def giveReport(frame, r):
        """
        Report of all the objects created
        frame => frame on which information is required to be rendered
        """
        WB = openpyxl.load_workbook("project.xlsx")
        products = WB["products"]
        MAXROW_admin = products.max_row
        for i in range(2, MAXROW_admin+1):
            tk.Label(frame, text=products[f"A{i}"].value, font=("Times New Roman", 13)).grid(row=r, column=0)
            tk.Label(frame, text=products[f"B{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=1)
            tk.Label(frame, text=products[f"C{i}"].value,font=("Times New Roman", 13)).grid(row=r, column=2)
            r = r + 1
        WB.save("project.xlsx")
        return r

    def isExist(id):
        WB = openpyxl.load_workbook("project.xlsx")
        products = WB["products"]
        MAXROW_product = products.max_row
        for i in range(1, MAXROW_product+1):
            if str(products[f"A{i}"].value) == str(id):
                return True
        WB.save("project.xlsx")
        return False
